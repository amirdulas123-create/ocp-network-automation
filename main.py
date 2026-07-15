import customtkinter as ctk          # Framework GUI (Tkinter stylise, dark/light mode)
import threading                     # Pour lancer les taches reseau sans geler l'UI
import os
import csv                           # Rapports CSV recapitulatifs (module stdlib, s'ouvre dans Excel)
import json
import re                            # Parsing des warnings/erreurs remontes par nmap
import time                          # Chronometrage du scan (temps total)
import shlex                         # Decoupage de la ligne de commande nmap (comme python-nmap)
import shutil                        # Localisation du binaire nmap (pour trouver nmap-services)
import functools                     # Cache du calcul des top ports UDP (lru_cache)
import subprocess                    # Utilise pour lancer "ping" systeme et nmap.exe
import ipaddress                     # Manipulation de sous-reseaux (ex: 192.168.1.0/24)
import socket                        # Test de ports ouverts (SSH/Telnet/HTTP/...)
import xml.etree.ElementTree as ET   # Parsing du XML Nmap (detection des hotes coupes par timeout)
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed  # Scan reseau en parallele
import tkinter as tk                 # TclError (widgets detruits) + acces bas niveau Tk
from tkinter import filedialog, PanedWindow  # Boite de dialogue "Parcourir un fichier" + separateur redimensionnable
# netmiko (SSH) et openpyxl (lecture Excel) sont importes PARESSEUSEMENT au moment
# ou on en a besoin (clic sur un bouton), pas au chargement du module : leur import
# coute ~0.9s a eux deux et rallongeait d'autant le demarrage de l'app avant meme
# l'affichage de la fenetre. nmap est deja importe paresseusement (voir plus bas).

HISTORY_FILE = os.path.join("data", "scan_history.json")   # Historique des scans reseau
SETTINGS_FILE = os.path.join("data", "settings.json")      # Preferences (theme dark/light)
COMMANDS_FILE = "commands.txt"                              # Commandes par defaut si la zone de texte est vide
REPORTS_DIR = "reports"                                     # Rapports CSV (un fichier par execution)
REPORT_HEADER = ["Date", "Appareil", "Operation", "Statut", "Detail"]
# Icone OCP partagee par la fenetre principale ET les fenetres de scan (voir _apply_icon).
ICON_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "ocp_logo.ico")
# Tuples (mode clair, mode sombre) explicites pour tout ce qui vit dans un PanedWindow :
# un PanedWindow est un widget Tk brut, donc la detection automatique de couleur de
# CustomTkinter ("transparent", coins arrondis) s'y resout UNE seule fois au demarrage
# et ne suit plus le changement de theme — d'ou bandes/coins de la mauvaise couleur.
# Avec des tuples explicites, CustomTkinter rebascule tout seul a chaque toggle.
# Fond du PanedWindow = meme couleur que le fond des onglets CTkTabview, pour que les
# coins arrondis des textbox et la zone du sash se fondent dans le decor au lieu de
# ressortir en gris plus fonce. Le sash reste reperable grace a sashrelief="raised".
PANED_BG = ("gray86", "gray17")
TAB_BG = ("gray86", "gray17")

def load_devices():
    # Lit devices.xlsx : 1ere ligne = headers, chaque ligne suivante = un appareil
    import openpyxl  # import paresseux (voir en-tete) : ~0.5s, inutile au demarrage
    wb = openpyxl.load_workbook("devices.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

def ping(ip):
    # Un seul ping ICMP (timeout 500ms) via la commande systeme Windows "ping"
    result = subprocess.run(
        ["ping", "-n", "1", "-w", "500", str(ip)],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
    )
    return str(ip), result.returncode == 0

def detect_device(ip):
    # Teste des ports TCP courants pour deviner le type d'appareil (pas d'auth, juste connect())
    ports = {22: "SSH", 23: "Telnet", 80: "HTTP", 443: "HTTPS", 3389: "RDP"}
    open_ports = []
    for port, name in ports.items():
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(0.3)
            if sock.connect_ex((ip, port)) == 0:  # 0 = port ouvert
                open_ports.append(name)
            sock.close()
        except:
            pass
    if "SSH" in open_ports or "Telnet" in open_ports:
        return "Equipement reseau"
    elif "RDP" in open_ports:
        return "PC Windows"
    elif "HTTP" in open_ports or "HTTPS" in open_ports:
        return "Serveur Web"
    elif open_ports:
        return f"Hote ({', '.join(open_ports)})"
    else:
        return "Hote actif"

class ScanCancelled(Exception):
    # Levee quand l'utilisateur demande l'arret du scan (bouton "Arreter le scan").
    # Remonte jusqu'a _run_scan_inner qui l'affiche proprement au lieu d'un "Erreur".
    pass
    # Case décochée = Nmap teste d'abord qui existe, puis scanne que ceux-là (plus rapide)
    # Case cochée = Nmap scanne tout le monde direct, sans vérifier avant (plus lent mais rate personne)
def scan_network(subnet, cancel=None):
    # Ping tous les hotes d'un sous-reseau en parallele (50 threads), puis detecte
    # le type de chaque appareil qui repond. cancel (threading.Event) permet d'arreter
    # entre deux hotes : on ne soumet plus de nouveau travail et on coupe l'attente.
    network = ipaddress.IPv4Network(subnet, strict=False)
    hosts = list(network.hosts())
    alive = []
    with ThreadPoolExecutor(max_workers=50) as executor:
        futures = {executor.submit(ping, ip): ip for ip in hosts}
        for future in as_completed(futures):
            if cancel is not None and cancel.is_set():
                # cancel_futures : n'attend pas les pings encore en file d'attente
                executor.shutdown(wait=False, cancel_futures=True)
                raise ScanCancelled()
            ip, is_alive = future.result()
            if is_alive:
                device_type = detect_device(ip)
                alive.append((ip, device_type))
    return sorted(alive, key=lambda x: x[0])

# ---------------------------------------------------------------------------
# Scan avance via Nmap (python-nmap)
# ---------------------------------------------------------------------------
# Ports TCP scannes par defaut : les memes que l'ancien scan basique
# (22/23/80/443/3389) + quelques ports pertinents en environnement Cisco/IT
# (FTP, NETCONF, consoles web alternatives). Modifiable via le champ "Ports"
# de l'onglet Scanner.
DEFAULT_SCAN_PORTS = "21,22,23,80,443,830,3389,8080,8443"
PORT_LABELS = {
    21: "FTP", 22: "SSH", 23: "Telnet", 80: "HTTP", 443: "HTTPS",
    830: "NETCONF", 3389: "RDP", 8080: "HTTP-alt", 8443: "HTTPS-alt",
    # Services UDP a risque (voir VULN_UDP_PORTS)
    53: "DNS", 67: "DHCP", 68: "DHCP", 69: "TFTP", 123: "NTP",
    161: "SNMP", 500: "ISAKMP", 1900: "SSDP",
}
# Le scan UDP N'EST PAS un inventaire exhaustif (scanner 65535 ports UDP est
# impraticable, surtout contre Windows). Objectif metier clarifie : TCP = inventaire
# reseau (appareils + services) ; UDP = RECHERCHE DE VULNERABILITES sur les services
# UDP les plus a risque en contexte Cisco/IT. Quand "Inclure UDP" est coche, la partie
# UDP scanne les N ports UDP les plus FREQUENTS (approche "top ports" de Nmap, basee sur
# les statistiques reelles du fichier nmap-services) -- meilleure couverture qu'une
# liste manuelle figee, pour un cout de temps similaire. Le champ "Ports" ne concerne
# que le TCP. Voir _udp_scan_ports().
#
# Compromis couverture/rapidite : plus de ports = meilleure couverture mais scan plus
# long. 100 est un bon point de depart (couvre l'essentiel des services UDP reels tout
# en restant praticable en quelques minutes).
TOP_UDP_PORTS_COUNT = 100
# Ports UDP a haut risque TOUJOURS inclus (meme s'ils sortaient du top-N), car ce sont
# les services que ciblent nos scripts NSE de vuln ci-dessous. Sert aussi de repli
# complet si nmap-services est introuvable (voir _udp_scan_ports).
VULN_UDP_PORTS = [
    161,   # SNMP  -- communaute par defaut (public), information disclosure
    69,    # TFTP  -- transfert de config sans authentification
    53,    # DNS   -- amplification, cache poisoning, resolveur ouvert
    123,   # NTP   -- amplification, monlist
    67,    # DHCP  (serveur) -- usurpation, rogue DHCP
    68,    # DHCP  (client)
    500,   # ISAKMP/IKE -- VPN souvent mal configure
    1900,  # SSDP/UPnP  -- souvent expose par erreur
]
# Scripts NSE de detection lances avec la partie UDP. Chaque script porte sa propre
# "portrule" limitee a SON service -> ils ne se declenchent que sur le service concerne
# (SNMP, TFTP, DNS...) quel que soit le port scanne, jamais sur la grande plage TCP
# d'inventaire. Elargir la couverture de ports (top-N) n'y change donc rien : la
# detection reste ciblee. On evite volontairement --script vuln (categorie entiere) :
# trop lente et bruyante, elle tournerait aussi sur tous les ports TCP.
VULN_UDP_SCRIPTS = ",".join([
    "snmp-info", "snmp-sysdescr",   # SNMP : infos systeme via communaute par defaut
    "tftp-enum",                     # TFTP : fichiers accessibles sans auth
    "dns-recursion",                 # DNS  : resolveur recursif ouvert (amplification)
    "ntp-info", "ntp-monlist",       # NTP  : infos + monlist (amplification)
    "dhcp-discover",                 # DHCP : reponse d'un serveur (rogue/usurpation)
    "ike-version",                   # ISAKMP/IKE : version/vendor du VPN
    "upnp-info",                     # SSDP/UPnP : services exposes
])

def _find_nmap_services():
    # Localise le fichier nmap-services (base de frequences des ports, celle qu'utilise
    # --top-ports). nm._nmap_path ne donne souvent que "nmap" -> on cherche via le PATH
    # puis dans les emplacements d'installation classiques (Windows / Linux / macOS).
    candidates = []
    exe = shutil.which("nmap")
    if exe:
        d = os.path.dirname(exe)
        candidates.append(os.path.join(d, "nmap-services"))                       # Windows : meme dossier
        candidates.append(os.path.join(d, "..", "share", "nmap", "nmap-services"))  # Unix : ../share/nmap
    candidates += [
        r"C:\Program Files (x86)\Nmap\nmap-services",
        r"C:\Program Files\Nmap\nmap-services",
        "/usr/share/nmap/nmap-services",
        "/usr/local/share/nmap/nmap-services",
        "/opt/homebrew/share/nmap/nmap-services",
    ]
    for c in candidates:
        if os.path.isfile(c):
            return c
    return None

@functools.lru_cache(maxsize=None)
def _top_udp_ports(n):
    # Renvoie les n ports UDP les plus frequents d'apres nmap-services (exactement la
    # base de --top-ports de Nmap : format "service port/proto frequence ..."). En repli
    # (fichier introuvable/illisible), on renvoie la liste curatee VULN_UDP_PORTS. Le
    # resultat est cache (lru_cache) : le fichier ~1 Mo n'est lu qu'une fois.
    path = _find_nmap_services()
    if not path:
        return tuple(VULN_UDP_PORTS)
    rows = []
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split()
                if len(parts) < 3:
                    continue
                try:
                    port_s, proto = parts[1].split("/")
                    freq = float(parts[2])
                except ValueError:
                    continue
                if proto == "udp":
                    rows.append((freq, int(port_s)))
    except OSError:
        return tuple(VULN_UDP_PORTS)
    if not rows:
        return tuple(VULN_UDP_PORTS)
    rows.sort(key=lambda x: x[0], reverse=True)
    return tuple(p for _, p in rows[:n])

def _udp_scan_ports():
    # Ports UDP effectivement scannes quand "Inclure UDP" est coche : union du top-N
    # (couverture) et des ports vuln curatee (garantit que les services cibles par nos
    # scripts NSE sont toujours scannes, meme si absents du top-N). Trie par numero pour
    # une chaine -p stable/lisible.
    return sorted(set(_top_udp_ports(TOP_UDP_PORTS_COUNT)) | set(VULN_UDP_PORTS))
# --- Modele de timeout par hote (--host-timeout) ---------------------------------
# time ≈ cout FIXE + cout marginal par port, avec marge de securite. Calibre sur des
# mesures reelles (cout fixe ~4-5s + ~0.0015 s/port TCP). La partie UDP (top-ports +
# scripts NSE de vuln) est bornee : cout sur udp_port_count + un cout fixe de scripts.
#
# IMPORTANT (cas "vulnerabilite reelle trouvee") : un script NSE qui NE trouve rien
# echoue vite (le service ne repond pas), mais un script qui TROUVE quelque chose doit
# interroger le service en detail (ex : SNMP walk pour sysdescr/uptime/interfaces quand
# la communaute "public" est active) -> il prend BEAUCOUP plus de temps. C'est pourtant
# le cas le plus important (detecter une vraie faille doit etre fiable). De plus, scanner
# 100 ports UDP sur un equipement Cisco reel est ralenti par le rate-limiting ICMP des
# reponses "port unreachable". Le cout fixe des scripts est donc dimensionne genereusement
# (150s) pour absorber PLUSIEURS scripts qui remontent vraiment des donnees + la lenteur
# UDP reelle, et non seulement le cas "rien ne repond" (qui, lui, finit vite). Un hote
# lent-mais-qui-repond (le plus precieux a scanner) ne doit jamais etre coupe avant la fin.
HOST_TIMEOUT_BASE = 30       # plancher (s) : couvre les ~9 ports par defaut
# Plafond : un hote qui repond en detail (vraie vuln) merite une large marge ; un hote
# muet, lui, finit tot de toute facon (aucune reponse a attendre) -> un plafond haut ne
# ralentit PAS le scan d'un /24 majoritairement "mort", il protege seulement le cas utile.
HOST_TIMEOUT_CAP = 2400      # plafond (s) : 40 min de securite (pire cas realiste ~12 min)
_FIXED_COST = 5.0            # cout fixe par phase de scan (s) : connexion + mise en place
_TCP_PER_PORT = 0.0015       # cout marginal TCP -sV (s/port) ; 65535 -> ~98s (mesure ~90s)
_UDP_PER_PORT = 0.005        # cout marginal UDP -sU (s/port), issu des mesures reelles
# Cout fixe des scripts NSE de vuln + de la lenteur du scan UDP sur equipement reel.
# 60s sous-estimait le cas "vuln reelle trouvee" (scan coupe a 251s sur un switch avec
# SNMP public actif) -> 150s pour couvrir plusieurs scripts qui repondent en detail.
_UDP_SCRIPT_COST = 150.0
# Cout de la detection d'OS (-O). Cout MARGINAL mesure ~3s sur un hote facile, MAIS Nmap
# relance le fingerprinting jusqu'a 5 fois sur un hote ambigu (equipement reseau Cisco)
# -> cout quasi FIXE mais pouvant monter a plusieurs dizaines de secondes. Valeur large.
_OS_DETECT_COST = 30.0
_TIMEOUT_SAFETY = 2.5        # marge de securite (x2.5, dans la fourchette x2-x3 voulue)

def _count_ports(ports):
    # Compte le nombre de ports decrits par une specification Nmap ("22,80",
    # "1-1024", "T:80,U:161", "1-65535"...). Sert a dimensionner --host-timeout.
    # Repli prudent sur 1 port par token illisible : on ne veut jamais renvoyer 0.
    total = 0
    for token in (ports or "").split(","):
        token = token.strip()
        if not token:
            continue
        if ":" in token:                    # prefixe protocole Nmap (T:, U:, S:)
            token = token.split(":", 1)[1]
        if "-" in token:
            try:
                a, b = token.split("-", 1)
                a = int(a) if a.strip() else 1
                b = int(b) if b.strip() else 65535
                total += max(1, b - a + 1)
            except ValueError:
                total += 1
        else:
            total += 1
    return total

def _host_timeout_for(ports, include_udp=False, os_detect=False, udp_port_count=0):
    # Timeout par hote (chaine "<n>s" pour Nmap), modele cout-fixe + cout-marginal.
    # Nmap peut executer jusqu'a trois phases : scan TCP -sV (sur le champ "Ports"),
    # scan UDP -sU + scripts (sur les udp_port_count ports "top", JAMAIS le champ
    # "Ports"), detection -O. On additionne leurs couts.
    #   9 ports TCP                    -> plancher 30s
    #   1-65535 TCP                    -> ~4-5 min
    #   9 ports TCP + UDP(100) + OS    -> ~8 min  (marge pour une vraie vuln qui repond)
    #   1-65535 TCP + UDP(100) + OS    -> ~12 min
    n = _count_ports(ports)                       # cote TCP uniquement (champ "Ports")
    est = _FIXED_COST + n * _TCP_PER_PORT
    if include_udp:
        # UDP borne : nombre de ports "top" (udp_port_count) + cout fixe des scripts NSE.
        # Le cout marginal par port UDP reste faible (~0.005s), le cout des scripts domine.
        est += _FIXED_COST + udp_port_count * _UDP_PER_PORT + _UDP_SCRIPT_COST
    if os_detect:
        est += _OS_DETECT_COST
    seconds = int(max(HOST_TIMEOUT_BASE, min(est * _TIMEOUT_SAFETY, HOST_TIMEOUT_CAP)))
    return f"{seconds}s"

def _has_admin_privileges():
    # Nmap exige des privileges admin/root (raw sockets) pour -O (detection d'OS) ET
    # pour -sU (scan UDP). On le detecte AVANT de lancer le scan pour ne lancer -O que
    # si c'est possible : cela evite la tentative -O vouee a l'echec suivie d'un
    # SECOND scan complet de repli (double du temps, cf. correction v1.5).
    try:
        if os.name == "nt":
            import ctypes
            return bool(ctypes.windll.shell32.IsUserAnAdmin())
        return hasattr(os, "geteuid") and os.geteuid() == 0
    except Exception:
        return False

def _timed_out_hosts(xml_output):
    # Nmap marque les hotes coupes par --host-timeout via l'attribut timedout="true"
    # sur l'element <host> du XML (-oX -). python-nmap n'expose PAS cet attribut, et
    # avec -oX - le message "Skipping host ..." n'apparait PAS sur stderr : sans ce
    # parsing, un hote coupe disparait silencieusement des resultats. On renvoie donc
    # l'ensemble des IP coupees pour pouvoir l'afficher clairement dans l'UI.
    hosts = set()
    if not xml_output:
        return hosts
    try:
        root = ET.fromstring(xml_output)
    except ET.ParseError:
        return hosts
    for host in root.findall("host"):
        if host.get("timedout") != "true":
            continue
        addr = None
        for a in host.findall("address"):
            if a.get("addrtype") in ("ipv4", "ipv6"):
                addr = a.get("addr")
                break
        hosts.add(addr or "?")
    return hosts

def _classify_device(open_ports):
    # Meme logique de classification que detect_device() (le fallback), mais a
    # partir des numeros de port ouverts remontes par Nmap. NETCONF (830) est
    # ajoute comme signe d'equipement reseau (frequent sur du Cisco recent / GNS3).
    if any(p in open_ports for p in (22, 23, 830)):
        return "Equipement reseau"
    if 3389 in open_ports:
        return "PC Windows"
    if any(p in open_ports for p in (80, 443, 8080, 8443)):
        return "Serveur Web"
    return "Hote actif"

def _best_os_match(host_info):
    # Renvoie la meilleure hypothese d'OS ("Linux 2.6.x (95%)") si -O a tourne,
    # sinon None. Nmap trie osmatch par precision decroissante -> [0] = meilleur.
    osmatch = host_info.get("osmatch", [])
    if not osmatch:
        return None
    best = osmatch[0]
    name = (best.get("name") or "").strip()
    if not name:
        return None
    accuracy = best.get("accuracy", "")
    return f"{name} ({accuracy}%)" if accuracy else name

def _format_nmap_host(host_info):
    # Construit la ligne de details d'un hote a partir des donnees Nmap, dans le
    # meme esprit que detect_device() mais enrichi : type d'equipement, services
    # avec produit/version (ex "SSH: OpenSSH 8.2"), et OS si disponible.
    services = []
    open_ports = []
    vuln_ids = []   # scripts NSE ayant produit une sortie (vuln potentielle)
    # TCP puis UDP (si un scan -sU a tourne). all_udp() existe dans python-nmap et
    # renvoie [] quand aucun port UDP n'a ete scanne, donc l'appel est sans risque.
    for proto in ("tcp", "udp"):
        all_ports = host_info.all_tcp() if proto == "tcp" else host_info.all_udp()
        for port in all_ports:
            pdata = host_info[proto][port]
            state = pdata.get("state")
            # Scripts NSE : on les remonte MEME si l'etat est "open|filtered" (frequent
            # en UDP), sinon on raterait une trouvaille de vuln sur un port non "open".
            for sid in (pdata.get("script") or {}):
                vuln_ids.append(f"{port}/{sid}")
            # En UDP, Nmap conclut souvent "open|filtered" (absence de reponse) : on
            # ne retient que les ports reellement "open" pour la liste des services.
            if state != "open":
                continue
            open_ports.append(port)
            label = PORT_LABELS.get(port, pdata.get("name") or str(port))
            if proto == "udp":
                label += "/udp"
            product = (pdata.get("product") or "").strip()
            version = (pdata.get("version") or "").strip()
            if product:
                services.append(f"{label}: {product}" + (f" {version}" if version else ""))
            else:
                services.append(label)
    detail = _classify_device(open_ports)
    if services:
        detail += " [" + ", ".join(services) + "]"
    os_label = _best_os_match(host_info)
    if os_label:
        detail += f" | OS: {os_label}"
    if vuln_ids:
        detail += " | ⚠ Vuln: " + ", ".join(sorted(set(vuln_ids)))
    return detail

def _shorten_script_output(output, limit=200):
    # Compacte une sortie de script NSE (souvent multi-lignes) en une ligne lisible
    # pour le log : espaces/retours ecrases, tronque au-dela de `limit` caracteres.
    compact = " | ".join(part.strip() for part in (output or "").splitlines() if part.strip())
    compact = re.sub(r"\s{2,}", " ", compact)
    return compact[:limit] + ("…" if len(compact) > limit else "")

def _host_script_findings(host_info):
    # Renvoie [(port_label, script_id, sortie_compacte), ...] pour tous les scripts NSE
    # (UDP en priorite) ayant produit une sortie : sert a lister les vulnerabilites
    # potentielles dans le log de l'UI. Ordre UDP d'abord (c'est la cible vuln).
    findings = []
    for proto in ("udp", "tcp"):
        all_ports = host_info.all_udp() if proto == "udp" else host_info.all_tcp()
        for port in all_ports:
            pdata = host_info[proto][port]
            scripts = pdata.get("script") or {}
            if not scripts:
                continue
            label = PORT_LABELS.get(port, pdata.get("name") or str(port))
            for sid, output in scripts.items():
                findings.append((f"{port}/{label}", sid, _shorten_script_output(output)))
    return findings

def _is_privilege_error(msg):
    # La detection d'OS (-O) exige les privileges admin/root. Sans eux, Nmap
    # s'arrete avec un message du genre "You requested a scan type which requires
    # root privileges." que python-nmap propage en PortScannerError.
    m = msg.lower()
    return "privile" in m or "requires root" in m or "requested a scan type" in m

def _ip_sort_key(item):
    # Tri numerique des IP (1.2.3.10 apres 1.2.3.9, pas l'inverse lexical).
    # Repli sur le tri texte si jamais Nmap remonte un nom d'hote.
    try:
        return (0, ipaddress.ip_address(item[0]))
    except ValueError:
        return (1, item[0])

def _host_responded(host_info):
    # On scanne TOUJOURS en -Pn (aucune decouverte), donc Nmap marque TOUTES les
    # IPs du sous-reseau comme "up". Un hote reellement present repond au moins par
    # un RST sur un port TCP ferme (ou ICMP port-unreachable en UDP, ou expose un
    # port ouvert) ; une IP morte ne renvoie que du 'filtered' ou rien. On ne garde
    # donc que les hotes qui ont vraiment repondu sur un port, pour ne pas lister
    # les 254 IPs du sous-reseau.
    for proto in ("tcp", "udp"):
        all_ports = host_info.all_tcp() if proto == "tcp" else host_info.all_udp()
        for port in all_ports:
            if host_info[proto][port].get("state") in ("open", "closed"):
                return True
    return False

def _run_nmap_process(nm, hosts, ports, arguments, cancel):
    # Replique PortScanner.scan() (python-nmap) mais via un Popen qu'on garde la main
    # dessus pour pouvoir le TERMINER sur demande d'annulation. nm.scan() est bloquant
    # et non interruptible : il lance nmap.exe et attend .communicate() jusqu'a la fin,
    # sans aucun point de controle. Ici on lance nmap.exe nous-memes (-oX - = XML sur
    # stdout), on draine stdout/stderr dans des threads (evite l'interblocage si le
    # buffer de pipe se remplit), et on sonde cancel toutes les 150ms pour pouvoir
    # .terminate() nmap.exe immediatement. Le XML recolte est ensuite parse par
    # python-nmap (analyse_nmap_xml_scan) pour rester 100% compatible avec nm[host].
    args = ([nm._nmap_path, "-oX", "-"] + shlex.split(hosts)
            + (["-p", ports] if ports else []) + shlex.split(arguments))
    proc = subprocess.Popen(args, bufsize=100000, stdin=subprocess.PIPE,
                            stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    out_holder, err_holder = {}, {}
    t_out = threading.Thread(target=lambda: out_holder.__setitem__("v", proc.stdout.read()), daemon=True)
    t_err = threading.Thread(target=lambda: err_holder.__setitem__("v", proc.stderr.read()), daemon=True)
    t_out.start()
    t_err.start()
    while proc.poll() is None:
        if cancel is not None and cancel.is_set():
            proc.terminate()  # TerminateProcess sur Windows : coupe nmap.exe pour de vrai
            try:
                proc.wait(timeout=3)
            except subprocess.TimeoutExpired:
                proc.kill()
            break
        time.sleep(0.15)
    t_out.join()
    t_err.join()
    if cancel is not None and cancel.is_set():
        raise ScanCancelled()
    xml_output = out_holder.get("v", b"")
    # errors="replace" : la stderr de Nmap (surtout avec --script) peut contenir des octets
    # non-UTF-8 (ex 0x92, apostrophe Windows-1252 dans un message de script) — un decode
    # strict planterait tout le scan. On tolere/remplace ces octets plutot que d'echouer.
    nmap_err = err_holder.get("v", b"").decode("utf-8", errors="replace")
    # Meme tri erreurs/warnings que PortScanner.scan() : un simple "Warning:" n'est pas
    # fatal, le reste de stderr l'est (fait lever PortScannerError par analyse_nmap_xml_scan).
    nmap_err_keep_trace, nmap_warn_keep_trace = [], []
    if len(nmap_err) > 0:
        regex_warning = re.compile("^Warning: .*", re.IGNORECASE)
        for line in nmap_err.split(os.linesep):
            if len(line) > 0:
                if regex_warning.search(line) is not None:
                    nmap_warn_keep_trace.append(line + os.linesep)
                else:
                    nmap_err_keep_trace.append(nmap_err)
    nm.analyse_nmap_xml_scan(nmap_xml_output=xml_output, nmap_err=nmap_err,
                             nmap_err_keep_trace=nmap_err_keep_trace,
                             nmap_warn_keep_trace=nmap_warn_keep_trace)
    # analyse_nmap_xml_scan a deja mis a jour nm._scan_result (utilise via nm.all_hosts()
    # par l'appelant). On renvoie en plus les hotes coupes par --host-timeout, invisibles
    # autrement (voir _timed_out_hosts), pour les signaler dans l'UI.
    return _timed_out_hosts(xml_output)


def _run_nmap_scan(nm, subnet, ports, log, include_udp=False, cancel=None):
    import nmap  # deja importe avec succes par l'appelant, juste pour PortScannerError
    # Decouverte d'hote : on force TOUJOURS -Pn (saute la phase de decouverte et
    # scanne directement les ports). Le benchmark (benchmark_scan.py) a montre que
    # -PE et -Pn donnent un temps quasi identique dans ce contexte (LAN/lab avec
    # --host-timeout court), tandis que -Pn garantit de ne jamais rater un hote qui
    # repond au ping mais que la decouverte Nmap par defaut (ARP/ICMP) ignore — cas
    # reel rencontre avec un routeur GNS3. On filtre ensuite les IPs "mortes" via
    # _host_responded() (en -Pn, Nmap marque toutes les IPs "up").
    # -T5 (au lieu de -T4) : timing le plus agressif, adapte a un lab/LAN local rapide
    #   ou la latence est faible et la perte de paquets negligeable. --min-hostgroup 64
    #   force Nmap a scanner les hotes par gros paquets (64 a la fois) au lieu de son
    #   decoupage par defaut plus petit : sur un /24 (254 hotes) ca augmente nettement
    #   le parallelisme inter-hotes et raccourcit le temps total (les 254 IPs sont
    #   toutes scannees en -Pn). --max-retries 1 coupe les retransmissions de sondes
    #   (inutiles en LAN) qui faisaient trainer les IPs sans reponse.
    # Detection des privileges AVANT le scan : on n'ajoute -O (detection d'OS, qui
    # exige les droits admin/root) que si on les a. Cela evite la sequence couteuse
    # "-O echoue -> on refait TOUT le scan sans -O" qui doublait le temps sur un gros
    # scan (mesure : 2054s ≈ 2x le timeout). Mesure confirmee : sans privileges, Nmap
    # -O echoue des le demarrage (raw sockets), il ne consomme donc pas de temps utile.
    # Il faut connaitre has_admin AVANT de calculer le timeout : -O ajoute un cout non
    # negligeable (fingerprinting + retries) que _host_timeout_for doit prendre en compte.
    has_admin = _has_admin_privileges()
    extra = " -O" if has_admin else ""
    # Ports scannes : le TCP (inventaire) suit le champ "Ports". L'UDP (recherche de vuln)
    # scanne les TOP_UDP_PORTS_COUNT ports UDP les plus frequents (union avec les ports
    # vuln curatee), quel que soit le champ "Ports" -> on les combine via la syntaxe Nmap
    # "T:<tcp>,U:<udp>". Il faut alors un type de scan TCP EXPLICITE (-sS) en plus de -sU,
    # sinon Nmap ignore la partie T: ("you haven't specified any TCP scan type"). NB :
    # --top-ports ne peut PAS coexister avec -p dans une seule invocation (verifie : -p
    # gagne, l'UDP est ignore), d'ou la liste de ports UDP calculee via nmap-services.
    # Les scripts NSE ne se declenchent que sur leurs services (portrules) -> detection
    # ciblee sur les services UDP, jamais sur la grande plage TCP.
    if include_udp:
        udp_ports = _udp_scan_ports()
        udp_ports_str = ",".join(str(p) for p in udp_ports)
        scan_ports = f"T:{ports},U:{udp_ports_str}"
        proto_args = f" -sS -sU --script {VULN_UDP_SCRIPTS}"
    else:
        udp_ports = []
        scan_ports = ports
        proto_args = ""
    # --host-timeout : modele cout-fixe + cout-marginal (voir _host_timeout_for). Le cout
    #   UDP est borne par le nombre de ports "top" scannes + le cout des scripts NSE.
    host_timeout = _host_timeout_for(ports, include_udp, os_detect=has_admin,
                                     udp_port_count=len(udp_ports))
    base_args = f"-sV{proto_args} -T5 --min-hostgroup 64 --max-retries 1 --host-timeout {host_timeout} -Pn"
    log("Decouverte d'hote : -Pn (aucune, scan direct des ports)")
    log(f"Scan TCP (inventaire) des ports : {ports}")
    if include_udp:
        log(f"Scan UDP (recherche de vulnerabilites) sur les {len(udp_ports)} ports UDP les "
            "plus courants (nmap-services) + scripts NSE (SNMP, TFTP, DNS, NTP, DHCP, ISAKMP, SSDP).")
    log(f"Timeout par hote : {host_timeout}"
        + (" (UDP top-ports + OS inclus)" if include_udp and has_admin
           else " (UDP top-ports inclus)" if include_udp
           else " (OS inclus)" if has_admin else "") + ".")
    if has_admin:
        log("Privileges admin detectes : detection d'OS (-O) activee.")
    else:
        log("Privileges admin absents : detection d'OS (-O) desactivee (versions de services conservees).")
        if include_udp:
            log("⚠ Le scan UDP (-sU) exige aussi les privileges admin : sans eux il echouera"
                " et l'app basculera sur le scan basique. Relancez l'app en administrateur.")
    try:
        timed_out = _run_nmap_process(nm, subnet, scan_ports, base_args + extra, cancel)
    except nmap.PortScannerError as e:
        # Filet de securite : si -O echoue malgre la detection de privileges (config
        # Npcap atypique), on relance UNE seule fois sans -O plutot que tout perdre.
        if extra and _is_privilege_error(str(e)):
            log("Detection d'OS (-O) a echoue malgre les privileges — nouvelle tentative sans -O.")
            timed_out = _run_nmap_process(nm, subnet, scan_ports, base_args, cancel)
        else:
            raise
    results = []
    for host in nm.all_hosts():
        host_info = nm[host]
        # En -Pn toutes les IPs sont "up" : on filtre sur une vraie reponse de port.
        if not _host_responded(host_info):
            continue
        results.append((host, _format_nmap_host(host_info)))
        # Vulnerabilites potentielles remontees par les scripts NSE : on les detaille
        # dans le log (le detail d'hote ne montre qu'un flag compact "⚠ Vuln: ...").
        findings = _host_script_findings(host_info)
        if findings:
            log(f"⚠ {host} — vulnerabilite(s) potentielle(s) detectee(s) :")
            for port_label, sid, summary in findings:
                log(f"    [{port_label}] {sid}" + (f" : {summary}" if summary else ""))
    found = sorted(results, key=_ip_sort_key)
    # Hotes coupes par --host-timeout : on les signale explicitement au lieu de les
    # laisser disparaitre en silence (le "0 hote detecte" serait alors trompeur).
    if timed_out:
        for ip in sorted(timed_out, key=lambda x: (_ip_sort_key((x, None)))):
            log(f"⚠ {ip} a depasse le timeout ({host_timeout}) — resultats potentiellement "
                "incomplets. Reduisez la plage de ports ou relancez pour ce seul hote.")
    log(f"{len(found)} hote(s) actif(s) detecte(s).")
    if not found and timed_out:
        log("Note : au moins un hote a repondu mais a ete coupe par le timeout avant de "
            "renvoyer un port exploitable — ce n'est pas une absence d'hote.")
    return found

def scan_network_smart(subnet, log, ports=DEFAULT_SCAN_PORTS, include_udp=False, cancel=None):
    # Point d'entree du scan : tente Nmap (detection OS + versions de services) et
    # retombe automatiquement sur scan_network() (ping + socket) si python-nmap ou
    # le binaire nmap manque, ou si le scan Nmap echoue. log(msg) ecrit dans l'UI.
    # cancel (threading.Event) est propage jusqu'a nmap.exe pour l'arret immediat.
    ports = (ports or DEFAULT_SCAN_PORTS).strip() or DEFAULT_SCAN_PORTS
    try:
        import nmap
    except ImportError:
        log("Nmap non disponible (module python-nmap absent) — scan basique utilise.")
        log("  Pour le scan avance : pip install python-nmap")
        return scan_network(subnet, cancel=cancel)
    try:
        nm = nmap.PortScanner()
    except nmap.PortScannerError:
        log("Nmap non disponible (binaire 'nmap' introuvable) — scan basique utilise.")
        log("  Installez Nmap : https://nmap.org/download.html puis pip install python-nmap")
        return scan_network(subnet, cancel=cancel)
    try:
        log("Scan Nmap en cours" + (" (TCP inventaire + UDP recherche de vulnerabilites)"
                                     if include_udp else " (TCP inventaire)") + "...")
        return _run_nmap_scan(nm, subnet, ports, log, include_udp=include_udp, cancel=cancel)
    except nmap.PortScannerError as e:
        log(f"Erreur Nmap : {e}")
        log("Repli sur le scan basique.")
        return scan_network(subnet, cancel=cancel)

def load_history():
    # Charge l'historique des scans (liste vide si le fichier n'existe pas encore)
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE) as f:
            return json.load(f)
    return []

def save_history(history):
    os.makedirs("data", exist_ok=True)
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2)

def load_settings():
    # Charge les preferences (theme). Dark par defaut si aucun fichier n'existe
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE) as f:
            return json.load(f)
    return {"theme": "dark"}

def save_settings(settings):
    os.makedirs("data", exist_ok=True)
    with open(SETTINGS_FILE, "w") as f:
        json.dump(settings, f, indent=2)

def report_row(appareil, operation, statut, detail):
    # Une ligne de rapport : Date,Appareil,Operation,Statut,Detail. Le detail est
    # aplati sur UNE seule ligne (les erreurs Netmiko et les sorties de switch sont
    # souvent multi-lignes) : sinon le module csv les entoure de guillemets et Excel
    # affiche une cellule qui deborde sur plusieurs lignes, illisible dans un tableau.
    detail = " ".join(str(detail).split())
    return [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), appareil, operation, statut, detail]

def save_report(prefix, rows):
    # Ecrit le recapitulatif d'une execution dans reports/<prefix>_<date>_<heure>.csv
    # et renvoie son chemin (None si l'operation n'a produit aucune ligne : rien a
    # rapporter, on ne cree pas de fichier vide).
    if not rows:
        return None
    os.makedirs(REPORTS_DIR, exist_ok=True)
    path = os.path.join(REPORTS_DIR, f"{prefix}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.csv")
    # newline="" : impose par le module csv, sinon Windows ajoute un \r en trop et le
    # fichier contient une ligne vide sur deux.
    # utf-8-sig = UTF-8 + BOM : c'est le BOM qui fait qu'Excel reconnait l'UTF-8 au
    # lieu de retomber sur l'encodage local (sans lui, "Equipement rejete" et les
    # accents des messages d'erreur s'affichent en "Ã©" dans les cellules).
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(REPORT_HEADER)
        writer.writerows(rows)
    return path

def theme_color(color_tuple):
    # tkinter.PanedWindow est un widget Tk brut (pas CustomTkinter) : il ne comprend
    # pas les tuples (couleur claire, couleur sombre) de CTk, il lui faut une couleur
    # resolue selon le mode d'apparence actuel.
    idx = 0 if ctk.get_appearance_mode() == "Light" else 1
    return color_tuple[idx]

class ScanPanel:
    """Panneau de scan reseau autonome et INDEPENDANT : ses propres champs (subnet,
    ports), bouton Scanner/Arreter, checkbox UDP, zone de log, et surtout son propre
    etat de scan (thread + threading.Event d'annulation). Aucun etat de scan n'est
    partage entre instances -> on peut en faire tourner plusieurs en parallele
    (onglet principal + fenetres 'Nouvelle fenetre de scan'), chacune menant son scan
    de bout en bout. Seul l'historique (scan_history.json) est commun a toutes les
    instances (fichier partage) ; son affichage n'existe que dans le panneau principal
    (show_history=True) et est rafraichi via App._refresh_history apres chaque scan."""

    def __init__(self, app, container, show_history=True, on_new_window=None):
        self.app = app                            # App principale : police partagee + refresh historique
        self._root = container.winfo_toplevel()   # fenetre hote (App ou CTkToplevel) pour .after() thread-safe
        self.show_history = show_history
        self._on_new_window = on_new_window
        self._scan_running = False
        self._scan_cancel = threading.Event()
        self._resize_after = None
        self._history_fit_bound = False
        self._build(container)

    # --- Construction de l'UI (identique a l'ancien _build_scanner) -----------------
    def _build(self, frame):
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(1, weight=1)
        # Barre du haut en GRILLE : chaque champ sur SA PROPRE LIGNE, largeur fixe, ancre
        # a gauche -> reste responsive sans deborder meme en fenetre etroite/zoomee.
        top = ctk.CTkFrame(frame, fg_color="transparent")
        top.grid(row=0, column=0, sticky="ew", pady=5)
        ctk.CTkLabel(top, text="Sous-reseau :").grid(row=0, column=0, padx=(5, 8), pady=(0, 6), sticky="w")
        self._subnet_entry = ctk.CTkEntry(top, placeholder_text="192.168.1.0/24", width=200)
        self._subnet_entry.grid(row=0, column=1, pady=(0, 6), sticky="w")
        # Ports optionnels : vide = DEFAULT_SCAN_PORTS. Syntaxe Nmap ("22,23,80", "1-1024").
        ctk.CTkLabel(top, text="Ports :").grid(row=1, column=0, padx=(5, 8), pady=(0, 6), sticky="w")
        self._ports_entry = ctk.CTkEntry(top, placeholder_text=DEFAULT_SCAN_PORTS, width=260)
        self._ports_entry.grid(row=1, column=1, pady=(0, 6), sticky="w")
        # Raccourci "Tous les ports" : remplit 1-65535 en un clic (--host-timeout s'adapte).
        ctk.CTkButton(top, text="Tous les ports", width=110,
                      command=self._fill_all_ports).grid(row=1, column=2, padx=(8, 0),
                                                          pady=(0, 6), sticky="w")
        # Bouton Scanner + checkbox UDP sur leur propre ligne (sous-frame packe a gauche).
        controls = ctk.CTkFrame(top, fg_color="transparent")
        controls.grid(row=2, column=0, columnspan=3, sticky="w", pady=(2, 0))
        # Largeur fixe : le texte passe de "Scanner" a "Arreter le scan" sans laisser de
        # contour fantome sur Windows (meme raison que le bouton d'envoi).
        self._scan_btn = ctk.CTkButton(controls, text="Scanner", width=130,
                                       fg_color="#dc2626", hover_color="#b91c1c",
                                       command=self._start_scan)
        self._scan_btn.pack(side="left", padx=(5, 10))
        # UDP = recherche de vulnerabilites (top-N ports UDP les plus courants + scripts
        # NSE), PAS un inventaire : le champ "Ports" ci-dessus ne concerne que le TCP.
        # Le libelle de la checkbox reste COURT (une CTkCheckBox ne sait pas passer son
        # texte a la ligne et le tronquait en fenetre etroite) ; tout le detail va dans le
        # label d'avertissement en dessous, lui responsive (wraplength, voir plus bas).
        self._udp_var = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(controls, text="Inclure UDP (-sU)",
                        variable=self._udp_var).pack(side="left")
        udp_hint = ctk.CTkLabel(
            top, justify="left",
            text=f"ⓘ Recherche de vulnerabilites sur les {TOP_UDP_PORTS_COUNT} ports UDP les plus "
                 "courants (SNMP, TFTP, DNS, NTP...) — independant du champ Ports (TCP), "
                 "compter 1-2 min de plus.",
            text_color="#d97706", font=ctk.CTkFont(size=11))
        udp_hint.grid(row=3, column=0, columnspan=3, sticky="ew", padx=5, pady=(2, 0))
        # Responsive : le texte revient a la ligne selon la largeur reelle de la fenetre
        # HOTE de ce panneau (principale OU fenetre de scan secondaire). _wrap_on_resize se
        # branche sur le <Configure> du parent du label (le bandeau "top" de CE panneau),
        # donc il suit bien la fenetre qui contient le panneau, pas forcement la principale.
        self.app._wrap_on_resize(udp_hint)
        # Bouton "Nouvelle fenetre de scan" : uniquement sur le panneau principal, pour
        # ouvrir une fenetre de scan supplementaire independante (voir App._open_scan_window).
        if self._on_new_window is not None:
            ctk.CTkButton(top, text="Nouvelle fenetre de scan", width=200,
                          fg_color="gray30", hover_color="gray40",
                          command=self._on_new_window).grid(
                              row=4, column=0, columnspan=3, sticky="w", padx=5, pady=(8, 0))

        # Log de scan (+ historique si panneau principal) dans un PanedWindow : sash a la
        # souris. opaqueresize=False + couleur en tuple pour suivre le theme (voir App).
        self._scan_paned = PanedWindow(frame, orient="vertical", sashwidth=6, sashrelief="raised",
                                        bg=theme_color(PANED_BG), bd=0, opaqueresize=False)
        self._scan_paned.grid(row=1, column=0, sticky="nsew", pady=(5, 0))
        # wrap="none" -> scrollbar horizontale interne : les longues lignes Nmap restent lisibles.
        self._scan_log = ctk.CTkTextbox(self._scan_paned, font=self.app._mono_font, wrap="none",
                                         state="disabled", bg_color=PANED_BG)
        self._scan_paned.add(self._scan_log, minsize=60, height=200)

        if self.show_history:
            history_pane = ctk.CTkFrame(self._scan_paned, fg_color=TAB_BG, corner_radius=0)
            history_pane.grid_columnconfigure(0, weight=1)
            history_pane.grid_rowconfigure(1, weight=1)
            ctk.CTkLabel(history_pane, text="Historique des scans",
                         font=ctk.CTkFont(size=13, weight="bold")).grid(row=0, column=0, sticky="w", pady=(5, 5))
            self._history_frame = ctk.CTkScrollableFrame(history_pane)
            self._history_frame.grid(row=1, column=0, sticky="nsew")
            self._history_frame.grid_columnconfigure(0, weight=1)
            self._scan_paned.add(history_pane, minsize=100, height=250)
            self._load_history_ui()
            self._install_history_resize_debounce()

    # --- Historique (panneau principal uniquement) ----------------------------------
    def refresh_history(self):
        # Rebranche l'affichage sur le thread principal ; ignore si le panneau n'a pas
        # d'historique (fenetre secondaire) ou a ete detruit entre-temps.
        if self.show_history:
            self._ui_after(self._load_history_ui)

    def _install_history_resize_debounce(self):
        # PERF : le CTkScrollableFrame de l'historique reflow tout son contenu a CHAQUE
        # <Configure> de son canvas -> ~3x plus lent en drag. On debranche ce fit pendant
        # le drag et on ne le rebranche qu'une fois la taille stabilisee (~160 ms).
        # Attributs internes CustomTkinter (5.2.2) : degrade proprement si absents.
        try:
            self._history_canvas = self._history_frame._parent_canvas
            self._history_fit = self._history_frame._fit_frame_dimensions_to_canvas
        except AttributeError:
            return
        self._history_fit_bound = True
        self._last_size = (self.app.winfo_width(), self.app.winfo_height())
        self.app.bind("<Configure>", self._on_window_configure, add="+")

    def _on_window_configure(self, event):
        # Ne reagit qu'au redimensionnement de la fenetre principale (App), pas aux
        # <Configure> d'autres widgets ni aux simples deplacements de fenetre.
        if event.widget is not self.app:
            return
        size = (event.width, event.height)
        if size == self._last_size:
            return
        self._last_size = size
        if self._history_fit_bound:
            self._history_canvas.unbind("<Configure>")
            self._history_fit_bound = False
        if self._resize_after is not None:
            self.app.after_cancel(self._resize_after)
        self._resize_after = self.app.after(160, self._resize_settled)

    def _resize_settled(self):
        self._resize_after = None
        if not self._history_canvas.winfo_exists():
            return
        self._history_canvas.bind("<Configure>", self._history_fit)
        self._history_fit_bound = True
        self._history_fit(None)
        self._history_canvas.configure(scrollregion=self._history_canvas.bbox("all"))

    def _load_history_ui(self):
        # Reconstruit l'affichage de l'historique depuis le JSON (repart de zero).
        if not self._history_frame.winfo_exists():
            return
        for widget in self._history_frame.winfo_children():
            widget.destroy()
        history = load_history()
        for i, entry in enumerate(history):
            row = ctk.CTkFrame(self._history_frame, fg_color="transparent")
            row.grid(row=i, column=0, sticky="ew", pady=2)
            row.grid_columnconfigure(1, weight=1)
            ctk.CTkLabel(row, text=entry["timestamp"], text_color="gray", width=160).grid(row=0, column=0, padx=5)
            lbl = ctk.CTkLabel(row, text=entry["target"], cursor="hand2")
            lbl.grid(row=0, column=1, sticky="w")
            # Clic sur le sous-reseau -> le remet dans le champ de saisie de CE panneau.
            lbl.bind("<Button-1>", lambda e, t=entry["target"]: (
                self._subnet_entry.delete(0, "end"),
                self._subnet_entry.insert(0, t)
            ))
            ctk.CTkButton(row, text="x", width=30, fg_color="transparent",
                          text_color=("gray20", "gray90"),
                          hover_color=("gray75", "gray30"),
                          command=lambda t=entry["target"]: self._delete_history(t)
                          ).grid(row=0, column=2, padx=5)

    def _delete_history(self, target):
        history = [h for h in load_history() if h["target"] != target]
        save_history(history)
        # Differe la reconstruction : _load_history_ui() detruit le bouton "x" en cours de clic.
        self._ui_after(self._load_history_ui)

    # --- Scan -----------------------------------------------------------------------
    def _fill_all_ports(self):
        # Remplace le contenu du champ Ports par la plage complete 1-65535 (syntaxe Nmap).
        self._ports_entry.delete(0, "end")
        self._ports_entry.insert(0, "1-65535")

    def request_cancel(self):
        # Appele a la fermeture de la fenetre : coupe le scan en cours. nmap.exe est
        # termine sous ~150ms via le polling de _scan_cancel dans _run_nmap_process.
        self._scan_cancel.set()

    def _start_scan(self):
        # Bouton start ET stop : pendant un scan, un nouveau clic ne relance pas de thread
        # (bug des resultats dupliques) — il demande l'arret via _scan_cancel. Le scan
        # Nmap etant un sous-processus bloquant, l'arret passe par la terminaison de
        # nmap.exe (voir _run_nmap_process), immediate et non differee a la fin du scan.
        if self._scan_running:
            self._scan_cancel.set()
            self._scan_btn.configure(state="disabled", text="Arret en cours...")
            return
        self._scan_running = True
        self._scan_cancel.clear()
        self._scan_btn.configure(text="Arreter le scan")
        threading.Thread(target=self._run_scan, daemon=True).start()

    def _run_scan(self):
        try:
            self._run_scan_inner()
        finally:
            self._scan_running = False
            self._ui_after(self._reset_scan_btn)

    def _reset_scan_btn(self):
        if self._scan_btn.winfo_exists():
            self._scan_btn.configure(state="normal", text="Scanner")

    def _run_scan_inner(self):
        # Scanne le sous-reseau saisi, affiche les resultats au fil de l'eau, sauvegarde
        # le rapport CSV et met a jour l'historique partage.
        self._clear()
        subnet = self._subnet_entry.get().strip()
        if not subnet:
            self._log("Entrez un sous-reseau valide")
            return
        self._log(f"Scan de {subnet} en cours...")
        t0 = time.perf_counter()
        try:
            ports = self._ports_entry.get().strip()
            include_udp = bool(self._udp_var.get())
            alive = scan_network_smart(subnet, self._log, ports=ports,
                                       include_udp=include_udp, cancel=self._scan_cancel)
            rows = []  # lignes du rapport CSV, une par hote trouve
            for ip, device_type in alive:
                self._log(f"✓ {ip} — {device_type}")
                rows.append(report_row(ip, "Scan", "Actif", device_type))
            self._log(f"\nTotal: {len(alive)} hotes actifs")
            self._log(f"Temps total du scan : {time.perf_counter() - t0:.1f}s")
            self._save_report("scan", rows)
            # Historique partage : retire l'ancienne entree du meme sous-reseau, la remet en tete.
            history = load_history()
            history = [h for h in history if h["target"] != subnet]
            history.insert(0, {"target": subnet, "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
            save_history(history)
            self.app._refresh_history()  # rafraichit l'affichage de l'onglet principal
        except ScanCancelled:
            # Arret demande (fermeture de fenetre ou bouton) : nmap.exe termine, pas d'historique.
            self._log(f"Scan annule par l'utilisateur (apres {time.perf_counter() - t0:.1f}s)")
        except Exception as e:
            self._log(f"Erreur: {e}")

    # --- Sortie (log/rapport), propres au panneau et thread-safe --------------------
    def _ui_after(self, func, *args):
        # Planifie func sur le thread principal Tk via la fenetre hote. Ignore silencieusement
        # si la fenetre a ete detruite (fermeture d'une fenetre de scan en cours).
        try:
            self._root.after(0, func, *args)
        except (tk.TclError, RuntimeError):
            pass

    def _log(self, msg):
        self._ui_after(self._log_ui, msg)

    def _log_ui(self, msg):
        box = self._scan_log
        if not box.winfo_exists():   # fenetre fermee pendant que le thread loggait encore
            return
        box.configure(state="normal")
        box.insert("end", msg + "\n")
        box.see("end")
        box.configure(state="disabled")

    def _clear(self):
        self._ui_after(self._clear_ui)

    def _clear_ui(self):
        box = self._scan_log
        if not box.winfo_exists():
            return
        box.configure(state="normal")
        box.delete("0.0", "end")
        box.configure(state="disabled")

    def _save_report(self, prefix, rows):
        # Chaque panneau genere son propre CSV (meme logique/prefixe que l'existant).
        try:
            path = save_report(prefix, rows)
        except OSError as e:
            self._log(f"Rapport non genere : {e}")
            return
        if path:
            self._log(f"Rapport genere : {path}")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        # CustomTkinter, sur Windows, cache (withdraw) puis reaffiche (deiconify) la
        # fenetre a CHAQUE appel de set_appearance_mode() rien que pour recolorer la
        # barre de titre (voir ctk_tk.py::_windows_set_titlebar_color) — c'est ca la
        # cause du "disappear/appear" au changement de theme, pas notre code. Ce flag
        # documente par CustomTkinter desactive cette manipulation de fenetre : la
        # barre de titre garde sa couleur par defaut, mais tout le contenu (boutons,
        # fond, texte) change de couleur instantanement, sans aucun clignotement.
        self._deactivate_windows_window_header_manipulation = True
        self._push_running = False
        self._push_cancel = threading.Event()
        # Meme protection anti-spam que l'envoi de config pour la sauvegarde : sans elle,
        # cliquer plusieurs fois lançait plusieurs sauvegardes en parallele (voir _start_backup).
        self._backup_running = False
        self._backup_cancel = threading.Event()
        # L'etat de scan (thread + annulation) vit desormais dans chaque ScanPanel : un
        # par onglet principal + un par fenetre 'Nouvelle fenetre de scan' (voir ScanPanel).
        self.settings = load_settings()
        ctk.set_appearance_mode(self.settings.get("theme", "dark"))
        self._zoom = self.settings.get("zoom", 1.0)
        ctk.set_widget_scaling(self._zoom)
        self.title("OCP Network Automation")
        self.geometry("900x700")
        self.minsize(700, 600)  # empeche de redimensionner trop petit (UI cassee sinon)
        # Icone de fenetre : purement decoratif. Un .ico manquant/corrompu ne doit jamais
        # empecher le demarrage -> _apply_icon est silencieux en cas d'echec.
        self._apply_icon(self)
        self._build()

    def _apply_icon(self, window):
        # Applique l'icone OCP a une fenetre (principale ou CTkToplevel). Sur une
        # CTkToplevel, CustomTkinter REPOSE sa propre icone ~200ms apres la creation
        # (voir ctk_toplevel), ecrasant un iconbitmap immediat -> on la reapplique aussi
        # en differe pour qu'elle "tienne". try/except : purement decoratif.
        def _set():
            try:
                window.iconbitmap(ICON_PATH)
            except Exception:
                pass
        _set()
        try:
            window.after(300, _set)
        except Exception:
            pass

    def _build(self):
        # Police mono partagee par toutes les zones de texte (commandes/logs) : une
        # seule CTkFont reutilisee au lieu d'en instancier une identique par textbox.
        self._mono_font = ctk.CTkFont(family="Consolas", size=11)

        # Layout principal : onglets (row 0) + bouton theme en bas (row 1)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=0)

        tabs = ctk.CTkTabview(self, anchor="nw")
        tabs.grid(row=0, column=0, padx=16, pady=(16,0), sticky="nsew")
        tabs.add("Envoi Config")
        tabs.add("Sauvegarde")
        tabs.add("Scanner Reseau")
        self._build_push(tabs.tab("Envoi Config"))
        self._build_backup(tabs.tab("Sauvegarde"))
        self._build_scanner(tabs.tab("Scanner Reseau"))

        # Barre du bas : zoom (petit, a gauche du bouton theme) + bouton theme
        bottom_bar = ctk.CTkFrame(self, fg_color="transparent")
        bottom_bar.grid(row=1, column=0, padx=16, pady=8, sticky="e")

        ctk.CTkButton(bottom_bar, text="-", width=28, fg_color="gray30", hover_color="gray40",
                      command=self._zoom_out).pack(side="left", padx=(0,2))
        self._zoom_label = ctk.CTkLabel(bottom_bar, text=f"{round(self._zoom * 100)}%", width=45)
        self._zoom_label.pack(side="left", padx=2)
        ctk.CTkButton(bottom_bar, text="+", width=28, fg_color="gray30", hover_color="gray40",
                      command=self._zoom_in).pack(side="left", padx=(2,10))

        # Le texte du bouton indique l'action a venir, pas l'etat actuel
        theme_label = "Mode clair" if self.settings.get("theme") == "dark" else "Mode sombre"
        self._theme_btn = ctk.CTkButton(bottom_bar, text=theme_label, width=120,
                                         fg_color="gray30", hover_color="gray40",
                                         command=self._toggle_theme)
        self._theme_btn.pack(side="left")

    def _zoom_in(self):
        self._set_zoom(self._zoom + 0.1)

    def _zoom_out(self):
        self._set_zoom(self._zoom - 0.1)

    def _set_zoom(self, zoom):
        # Bornes raisonnables pour eviter un texte illisible (trop petit) ou une
        # fenetre qui deborde de l'ecran (trop grand)
        zoom = max(0.7, min(1.5, round(zoom, 2)))
        self._zoom = zoom
        ctk.set_widget_scaling(zoom)
        self._zoom_label.configure(text=f"{round(zoom * 100)}%")
        self.settings["zoom"] = zoom
        save_settings(self.settings)
        # La fenetre garde TOUJOURS la taille choisie par l'utilisateur (resize manuel
        # ou taille par defaut) — le zoom ne doit jamais y toucher. Si le contenu
        # grossi deborde de la fenetre existante, c'est aux widgets d'etre scrollables
        # (voir _build_push et _build_scanner), pas a la fenetre de s'agrandir.

    def _toggle_theme(self):
        # La cause du glitch (voir __init__) est desactivee via
        # _deactivate_windows_window_header_manipulation : plus besoin de bidouiller
        # la fenetre ici, set_appearance_mode() se contente de recolorer les widgets.
        current = self.settings.get("theme", "dark")
        new_theme = "light" if current == "dark" else "dark"
        self.settings["theme"] = new_theme
        save_settings(self.settings)

        ctk.set_appearance_mode(new_theme)
        self._theme_btn.configure(text="Mode clair" if new_theme == "dark" else "Mode sombre")

        # Les PanedWindow sont des widgets Tk bruts : leur couleur ne suit pas
        # set_appearance_mode(), il faut la reappliquer nous-memes. Le scan a desormais
        # un PanedWindow par ScanPanel (onglet principal + chaque fenetre de scan ouverte).
        paned_bg = theme_color(PANED_BG)
        paneds = [getattr(self, "_push_paned", None)]
        main_panel = getattr(self, "_main_scan_panel", None)
        if main_panel is not None:
            paneds.append(main_panel._scan_paned)
        for _win, panel in getattr(self, "_scan_windows", []):
            paneds.append(getattr(panel, "_scan_paned", None))
        for paned in paneds:
            if paned is not None:
                try:
                    paned.configure(bg=paned_bg)
                except tk.TclError:
                    pass

    def _wrap_on_resize(self, label, pad=28):
        # Fait que le texte du label REVIENNE A LA LIGNE selon la largeur reelle
        # disponible, au lieu d'etre tronque quand la fenetre est etroite (surtout a
        # fort zoom). On suit le <Configure> du conteneur du label et on ajuste son
        # wraplength. Le garde-fou "si la valeur change vraiment" evite de reconfigurer
        # le label a chaque pixel -> cout negligeable pendant un resize.
        #
        # ATTENTION MISE A L'ECHELLE : event.width est en pixels ECRAN REELS, mais
        # CTkLabel interprete wraplength en unites LOGIQUES qu'il remultiplie par le
        # facteur d'echelle (zoom x DPI). Passer directement des pixels reels -> double
        # mise a l'echelle -> wraplength ~2x trop grand -> le texte ne revenait JAMAIS a
        # la ligne a fort zoom (bug visible dans la fenetre de scan etroite). On divise
        # donc par le facteur d'echelle du widget. A l'echelle 1.0, comportement inchange.
        parent = label.master
        def _update(event):
            sf = ctk.ScalingTracker.get_widget_scaling(label) or 1.0
            wl = max(120, int(event.width / sf) - pad)
            if label.cget("wraplength") != wl:
                label.configure(wraplength=wl)
        parent.bind("<Configure>", _update, add="+")

    def _build_push(self, frame):
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(3, weight=1)

        # Header
        header = ctk.CTkFrame(frame, fg_color="transparent")
        header.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(5,10))
        header.grid_columnconfigure(0, weight=1)
        title = ctk.CTkLabel(header, text="Envoi de Configuration Multi-Appareils",
                             font=ctk.CTkFont(size=15, weight="bold"), justify="left")
        title.grid(row=0, column=0, sticky="w")
        # justify="left" + wraplength responsive : le texte passe a la ligne au lieu
        # d'etre coupe en fenetre etroite/zoomee (voir _wrap_on_resize).
        desc = ctk.CTkLabel(header, text="Tapez des commandes ou chargez un fichier .txt — vide = commands.txt utilise",
                            text_color="gray", justify="left")
        desc.grid(row=1, column=0, sticky="w")
        self._wrap_on_resize(title)
        self._wrap_on_resize(desc)

        # Buttons row
        btn_frame = ctk.CTkFrame(frame, fg_color="transparent")
        btn_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0,5))
        # Largeur fixe : le texte change ("Envoyer a tous les appareils" -> "Envoi en
        # cours...") pendant l'envoi, et sans largeur fixe le canvas interne du bouton
        # se redimensionne pour suivre le texte plus court, ce qui laisse un contour
        # fantome de l'ancienne taille le temps d'un repaint sur Windows (glitch visuel).
        self._push_btn = ctk.CTkButton(btn_frame, text="Envoyer a tous les appareils",
                      width=220,
                      fg_color="#1d4ed8", hover_color="#1e40af",
                      command=self._start_push)
        self._push_btn.pack(side="left", padx=(0,10))
        ctk.CTkButton(btn_frame, text="Parcourir .txt",
                      fg_color="gray30", hover_color="gray40",
                      command=self._browse_commands
                      ).pack(side="left")

        # Commands label
        ctk.CTkLabel(frame, text="Commandes :").grid(row=2, column=0, sticky="w", pady=(5,0))

        # Zone de commandes + log dans un PanedWindow : le sash (barre entre les deux)
        # se tire a la souris pour agrandir/reduire chaque zone, au lieu d'une hauteur
        # fixe ou d'un partage automatique qui ne convient jamais a tout le monde.
        # opaqueresize=False : le redimensionnement ne s'applique qu'au relachement de
        # la souris (une ligne fantome pendant le drag) — le resize continu pendant le
        # drag faisait glitcher les widgets CustomTkinter contenus dans les panes.
        self._push_paned = PanedWindow(frame, orient="vertical", sashwidth=6, sashrelief="raised",
                                        bg=theme_color(PANED_BG), bd=0, opaqueresize=False)
        self._push_paned.grid(row=3, column=0, columnspan=2, sticky="nsew", pady=(2,5))

        # Commands text area (editable, une commande par ligne)
        # bg_color=PANED_BG en tuple : la couleur derriere les coins arrondis suit le
        # theme automatiquement (voir commentaire de PANED_BG en haut du fichier)
        self._commands_box = ctk.CTkTextbox(self._push_paned, font=self._mono_font,
                                             bg_color=PANED_BG)
        self._push_paned.add(self._commands_box, minsize=60, height=150)

        # Log output (lecture seule, resultat de l'envoi)
        # wrap="none" : pas de retour a la ligne automatique -> les lignes longues
        # (ex erreurs Netmiko) restent sur une seule ligne. CTkTextbox affiche alors
        # tout seul sa scrollbar horizontale interne (_x_scrollbar, gere par
        # _check_if_scrollbars_needed) quand une ligne deborde, et la masque sinon.
        self._push_log = ctk.CTkTextbox(self._push_paned, font=self._mono_font, wrap="none",
                                         state="disabled", bg_color=PANED_BG)
        self._push_paned.add(self._push_log, minsize=60, height=250)

    def _browse_commands(self):
        # Charge un .txt et remplace le contenu de la zone de texte des commandes
        filepath = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if filepath:
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read()
            self._commands_box.delete("0.0", "end")
            self._commands_box.insert("0.0", content)

    def _build_backup(self, frame):
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(3, weight=1)

        # Titre, description et bouton chacun sur SA propre ligne (ancres a gauche) :
        # avant, le titre et le bouton partageaient la ligne 0 et, en fenetre etroite
        # (ou a fort zoom), le bouton sans weight prenait toute sa place et repoussait /
        # tronquait le titre. En les empilant (comme la description l'etait deja), plus
        # aucun element n'est en concurrence de largeur ni pousse hors de l'ecran.
        title = ctk.CTkLabel(frame, text="Sauvegarde de Configuration",
                             font=ctk.CTkFont(size=15, weight="bold"), justify="left")
        title.grid(row=0, column=0, sticky="w", pady=(5,0))
        # wraplength responsive : le texte revient a la ligne au lieu d'etre tronque
        # en fenetre etroite/zoomee (voir _wrap_on_resize).
        desc = ctk.CTkLabel(frame, text="Sauvegarde la config active de tous les appareils dans des fichiers horodates",
                            text_color="gray", justify="left")
        desc.grid(row=1, column=0, sticky="w")
        self._wrap_on_resize(title)
        self._wrap_on_resize(desc)
        # Largeur fixe (comme _push_btn/_scan_btn) : le texte passe de "Sauvegarder tous
        # les appareils" a "Arreter la sauvegarde" ; sans largeur fixe, le retrecissement
        # laisse un contour fantome sur Windows. command=_start_backup (pas un thread
        # direct) pour la protection anti-spam start/stop.
        self._backup_btn = ctk.CTkButton(frame, text="Sauvegarder tous les appareils",
                      width=240, fg_color="#15803d", hover_color="#166534",
                      command=self._start_backup)
        self._backup_btn.grid(row=2, column=0, pady=(8,5), sticky="w")

        # wrap="none" -> scrollbar horizontale interne auto quand une ligne deborde (voir _build_push)
        self._backup_log = ctk.CTkTextbox(frame, font=self._mono_font, wrap="none", state="disabled")
        self._backup_log.grid(row=3, column=0, sticky="nsew", pady=10)

    def _build_scanner(self, frame):
        # L'onglet principal heberge un ScanPanel complet (avec l'historique partage) et
        # un bouton pour ouvrir des fenetres de scan supplementaires, chacune totalement
        # independante (voir ScanPanel / _open_scan_window).
        self._scan_windows = []
        self._main_scan_panel = ScanPanel(self, frame, show_history=True,
                                          on_new_window=self._open_scan_window)

    def _open_scan_window(self):
        # Ouvre une fenetre de scan independante (CTkToplevel) avec son PROPRE ScanPanel :
        # champs, log, thread et annulation geres de A a Z, sans etat partage. Permet de
        # scanner plusieurs sous-reseaux en parallele sans attendre la fin du premier.
        win = ctk.CTkToplevel(self)
        win.title("Scan reseau")
        win.geometry("640x600")
        win.minsize(560, 480)
        self._apply_icon(win)  # meme logo OCP que la fenetre principale (reapplique en differe)
        win.grid_columnconfigure(0, weight=1)
        win.grid_rowconfigure(0, weight=1)
        container = ctk.CTkFrame(win, fg_color="transparent")
        container.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
        panel = ScanPanel(self, container, show_history=False)
        self._scan_windows.append((win, panel))
        # Fermeture de la fenetre = annulation propre du scan en cours (voir _close_scan_window).
        win.protocol("WM_DELETE_WINDOW", lambda: self._close_scan_window(win, panel))
        # Premier plan + focus : une CTkToplevel s'ouvre sinon souvent DERRIERE la fenetre
        # principale (l'utilisateur doit la chercher). On la remonte (lift) et on la passe
        # brievement topmost pour forcer l'affichage devant, puis on relache le topmost pour
        # ne pas la coller au-dessus de tout en permanence. focus_force lui donne le clavier.
        win.lift()
        win.focus_force()
        win.attributes("-topmost", True)
        win.after(300, lambda: win.winfo_exists() and win.attributes("-topmost", False))

    def _close_scan_window(self, win, panel):
        # Ferme une fenetre de scan en ANNULANT d'abord son scan (arret de nmap.exe via son
        # threading.Event), pour ne pas laisser un sous-processus tourner sans controle.
        panel.request_cancel()
        self._scan_windows = [(w, p) for (w, p) in self._scan_windows if w is not win]
        win.destroy()

    def _refresh_history(self):
        # Rafraichit l'affichage de l'historique de l'onglet principal (fichier commun a
        # toutes les fenetres). Appele depuis le thread de n'importe quel ScanPanel.
        self._main_scan_panel.refresh_history()

    def _start_push(self):
        # Le bouton fait office de start ET stop : pendant l'envoi, un nouveau clic
        # ne relance pas un thread (ce qui recreerait le bug des clics multiples) —
        # il demande juste l'arret via un threading.Event verifie entre deux appareils.
        if self._push_running:
            self._push_cancel.set()
            self._push_btn.configure(state="disabled", text="Arret en cours...")
            return
        self._push_running = True
        self._push_cancel.clear()
        self._push_btn.configure(fg_color="#dc2626", hover_color="#b91c1c", text="Arreter l'envoi")
        threading.Thread(target=self._run_push, daemon=True).start()

    def _run_push(self):
        try:
            self._run_push_inner()
        finally:
            self._push_running = False
            self.after(0, lambda: self._push_btn.configure(
                state="normal", fg_color="#1d4ed8", hover_color="#1e40af",
                text="Envoyer a tous les appareils"))

    def _run_push_inner(self):
        # Envoie les commandes de configuration a tous les appareils de devices.xlsx
        from netmiko import ConnectHandler  # import paresseux (voir en-tete) : ~0.4s
        box = self._push_log
        self._clear(box)
        devices = load_devices()

        # Get commands from text area or fallback to commands.txt
        text_content = self._commands_box.get("0.0", "end").strip()
        if text_content:
            commands = [line.strip() for line in text_content.splitlines() if line.strip()]
        elif os.path.exists(COMMANDS_FILE):
            with open(COMMANDS_FILE, "r", encoding="utf-8") as f:
                commands = [line.strip() for line in f if line.strip()]
        else:
            self._log(box, "Aucune commande trouvee — zone de texte vide et commands.txt introuvable")
            return

        # commands.txt peut exister mais etre vide (ou ne contenir que des lignes vides) :
        # il faut verifier le CONTENU, pas juste la presence du fichier, sinon on continue
        # avec une liste vide -> connexion aux appareils pour "envoyer" 0 commande et un
        # faux message de succes ("config appliquee") alors que rien n'a ete fait.
        if not commands:
            self._log(box, "Aucune commande trouvee — zone de texte et commands.txt sont vides")
            return

        self._log(box, f"Commandes a envoyer: {len(commands)}")
        rows = []  # lignes du rapport CSV, une par appareil traite
        for d in devices:
            # Verifie AVANT chaque appareil si l'utilisateur a demande l'arret (bouton
            # "Arreter l'envoi"). On ne peut pas interrompre une connexion SSH deja en
            # cours au milieu d'un send_config_set, mais on evite au moins de continuer
            # sur les appareils suivants.
            if self._push_cancel.is_set():
                self._log(box, "Envoi arrete par l'utilisateur")
                break  # (et non return) : les appareils deja traites restent rapportes
            self._log(box, f"Connexion a {d['name']}...")
            try:
                conn = ConnectHandler(
                    device_type=d["device_type"], host=d["host"],
                    port=int(d["port"]), username=d["username"],
                    password=d["password"], secret=d["secret"],
                    # Desactive certains algos SSH modernes non supportes par les vieux Cisco
                    disabled_algorithms=dict(
                        kex=["curve25519-sha256", "curve25519-sha256@libssh.org",
                             "ecdh-sha2-nistp256", "ecdh-sha2-nistp384", "ecdh-sha2-nistp521",
                             "diffie-hellman-group16-sha512", "diffie-hellman-group-exchange-sha256"],
                        pubkeys=["rsa-sha2-512", "rsa-sha2-256"]
                    )
                )
                conn.enable()                     # mode privilegie
                output = conn.send_config_set(commands)  # config terminal + envoi des commandes
                conn.save_config()                # equivalent "write memory"
                conn.disconnect()

                # Le switch peut repondre "% Invalid input" / "% Incomplete command" etc.
                # sans que Netmiko leve d'exception -> on verifie nous-memes la sortie
                error_markers = ["% Invalid input", "% Incomplete command",
                                  "% Ambiguous command", "% Unrecognized command"]
                errors_found = [line for line in output.splitlines()
                                 if any(marker in line for marker in error_markers)]
                if errors_found:
                    self._log(box, f"⚠ {d['name']} — commande(s) rejetee(s) par le switch:")
                    for line in errors_found:
                        self._log(box, f"   {line.strip()}")
                    rows.append(report_row(d["name"], "Config", "Commande rejetee",
                                           " | ".join(line.strip() for line in errors_found)))
                else:
                    self._log(box, f"✓ {d['name']} — config appliquee")
                    rows.append(report_row(d["name"], "Config", "Succes", " | ".join(commands)))
            except Exception as e:
                # Continue sur les appareils suivants meme si celui-ci echoue
                self._log(box, f"✗ {d['name']} — {e}")
                rows.append(report_row(d["name"], "Config", "Echec", e))

        self._save_report(box, "config", rows)

    def _start_backup(self):
        # Meme pattern start/stop que _start_push : pendant une sauvegarde, un nouveau clic
        # ne relance PAS un thread (ce qui lançait plusieurs sauvegardes en parallele) — il
        # demande l'arret via _backup_cancel, verifie entre deux appareils.
        if self._backup_running:
            self._backup_cancel.set()
            self._backup_btn.configure(state="disabled", text="Arret en cours...")
            return
        self._backup_running = True
        self._backup_cancel.clear()
        self._backup_btn.configure(fg_color="#dc2626", hover_color="#b91c1c",
                                   text="Arreter la sauvegarde")
        threading.Thread(target=self._run_backup, daemon=True).start()

    def _run_backup(self):
        try:
            self._run_backup_inner()
        finally:
            self._backup_running = False
            self.after(0, lambda: self._backup_btn.configure(
                state="normal", fg_color="#15803d", hover_color="#166534",
                text="Sauvegarder tous les appareils"))

    def _run_backup_inner(self):
        # Recupere show running-config de chaque appareil et le sauvegarde dans un fichier horodate
        from netmiko import ConnectHandler  # import paresseux (voir en-tete) : ~0.4s
        box = self._backup_log
        self._clear(box)
        devices = load_devices()
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        backup_dir = os.path.join("backups", timestamp)  # un dossier par session de sauvegarde
        os.makedirs(backup_dir, exist_ok=True)
        rows = []  # lignes du rapport CSV, une par appareil
        for d in devices:
            # Verifie AVANT chaque appareil si l'arret a ete demande (bouton "Arreter la
            # sauvegarde"). On n'interrompt pas une connexion SSH deja en cours, mais on
            # evite de continuer sur les appareils suivants.
            if self._backup_cancel.is_set():
                self._log(box, "Sauvegarde arretee par l'utilisateur")
                break  # (et non return) : les appareils deja sauvegardes restent rapportes
            self._log(box, f"Connexion a {d['name']}...")
            try:
                conn = ConnectHandler(
                    device_type=d["device_type"], host=d["host"],
                    port=int(d["port"]), username=d["username"],
                    password=d["password"], secret=d["secret"],
                    disabled_algorithms=dict(
                        kex=["curve25519-sha256", "curve25519-sha256@libssh.org",
                             "ecdh-sha2-nistp256", "ecdh-sha2-nistp384", "ecdh-sha2-nistp521",
                             "diffie-hellman-group16-sha512", "diffie-hellman-group-exchange-sha256"],
                        pubkeys=["rsa-sha2-512", "rsa-sha2-256"]
                    )
                )
                conn.enable()
                output = conn.send_command("show running-config")
                conn.disconnect()
                filename = os.path.join(backup_dir, f"{d['name']}_{timestamp}.txt")
                with open(filename, "w") as f:
                    f.write(output)
                self._log(box, f"✓ {d['name']} — sauvegarde: {filename}")
                rows.append(report_row(d["name"], "Backup", "Succes", filename))
            except Exception as e:
                self._log(box, f"✗ {d['name']} — {e}")
                rows.append(report_row(d["name"], "Backup", "Echec", e))

        self._save_report(box, "backup", rows)

    def _save_report(self, box, prefix, rows):
        # Ecrit le CSV de l'operation qui vient de se terminer et annonce son chemin
        # dans le log de l'onglet : le log de l'UI disparait a la fermeture de l'app,
        # le rapport est ce qui reste. Un souci d'ecriture (dossier en lecture seule,
        # fichier ouvert dans Excel) est signale mais n'invalide pas l'operation, qui
        # elle a bien eu lieu -> on log l'erreur au lieu de la laisser remonter.
        try:
            path = save_report(prefix, rows)
        except OSError as e:
            self._log(box, f"Rapport non genere : {e}")
            return
        if path:
            self._log(box, f"Rapport genere : {path}")

    def _log(self, box, msg):
        # _run_push_inner/_run_backup/_run_scan tournent dans des threads en arriere-
        # plan, mais Tkinter n'est PAS thread-safe : manipuler des widgets (et surtout
        # appeler .update(), qui force le traitement de TOUTE la queue d'evenements Tk)
        # depuis un thread autre que le principal course avec les redraws internes de
        # CustomTkinter et provoque des TclError aleatoires ("invalid command name"),
        # surtout quand plusieurs operations tournent en meme temps. self.after(0, ..)
        # est le mecanisme thread-safe standard pour renvoyer le travail au thread
        # principal (fonctionne aussi si _log est deja appele depuis le thread principal).
        self.after(0, self._log_ui, box, msg)

    def _log_ui(self, box, msg):
        # Ecrit dans une textbox en lecture seule : deverrouille, ecrit, reverrouille, scroll bas
        box.configure(state="normal")
        box.insert("end", msg + "\n")
        box.see("end")
        box.configure(state="disabled")

    def _clear(self, box):
        self.after(0, self._clear_ui, box)

    def _clear_ui(self, box):
        box.configure(state="normal")
        box.delete("0.0", "end")
        box.configure(state="disabled")

if __name__ == "__main__":
    app = App()
    app.mainloop()